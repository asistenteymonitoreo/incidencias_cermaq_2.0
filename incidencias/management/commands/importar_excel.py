# incidencias/management/commands/importar_excel.py
"""
Comando para importar incidencias desde el Excel.

Uso:
    # Solo ver qué se importaría (no escribe nada):
    python manage.py importar_excel --dry-run

    # Importar solo las primeras 5 filas (para probar):
    python manage.py importar_excel --limit 5

    # Importar todo:
    python manage.py importar_excel
"""
import os
from datetime import datetime, time

from django.core.management.base import BaseCommand
from django.utils.text import slugify
from django.conf import settings

from openpyxl import load_workbook

from incidencias.models import Centro, Operario, Incidencia


class Command(BaseCommand):
    help = "Importa incidencias desde 'Reporte de Incidencias DIA COMPLETO.xlsm'"

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Solo muestra qué se importaría, sin escribir en la base de datos',
        )
        parser.add_argument(
            '--limit',
            type=int,
            default=0,
            help='Limitar a N filas (0 = sin límite)',
        )

    def handle(self, *args, **options):
        dry_run = options['dry_run']
        limit = options['limit']

        if dry_run:
            self.stdout.write(self.style.WARNING("=== MODO DRY-RUN: No se escribirá nada ===\n"))

        # Ruta del archivo
        file_path = os.path.join(settings.BASE_DIR, "Reporte de Incidencias DIA COMPLETO.xlsm")

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"Archivo no encontrado: {file_path}"))
            return

        self.stdout.write(f"Abriendo archivo: {file_path}")
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb["BASE DE DATOS"]

        self.stdout.write(f"Filas totales: {ws.max_row}, Columnas: {ws.max_column}")

        # Mapeo de columnas (basado en fila 7)
        # Col 1: Fecha, Col 2: Hora, Col 3: Turno, Col 4: Centro, Col 5: Módulo,
        # Col 6: Estanque, Col 7: TipoIncidencia, Col 8: Sensor, Col 9: Proveedor,
        # Col 10: Valor, Col 11: TiempoResolucion, Col 12: RiesgoPeces,
        # Col 13: PerdidaEconomica, Col 14: RiesgoPersonas, Col 15: Observaciones,
        # Col 16: Persona llamada, Col 17: Cargo, Col 18: Observacion Llamada

        HEADER_ROW = 7
        DATA_START_ROW = 8

        # Contadores
        centros_creados = 0
        operarios_creados = 0
        incidencias_creadas = 0
        filas_procesadas = 0
        filas_vacias = 0

        # Determinar última fila a procesar
        end_row = ws.max_row + 1
        if limit > 0:
            end_row = min(DATA_START_ROW + limit, ws.max_row + 1)
            self.stdout.write(f"Limitando a {limit} filas (filas {DATA_START_ROW} a {end_row - 1})")

        self.stdout.write("\n" + "=" * 60)

        for row_num in range(DATA_START_ROW, end_row):
            # Leer valores de la fila
            fecha = ws.cell(row=row_num, column=1).value
            hora = ws.cell(row=row_num, column=2).value
            turno = ws.cell(row=row_num, column=3).value
            centro_nombre = ws.cell(row=row_num, column=4).value
            modulo = ws.cell(row=row_num, column=5).value
            estanque = ws.cell(row=row_num, column=6).value
            tipo_incidencia = ws.cell(row=row_num, column=7).value
            sensor = ws.cell(row=row_num, column=8).value
            proveedor = ws.cell(row=row_num, column=9).value
            valor = ws.cell(row=row_num, column=10).value
            tiempo_resolucion = ws.cell(row=row_num, column=11).value
            riesgo_peces_str = ws.cell(row=row_num, column=12).value
            perdida_economica_str = ws.cell(row=row_num, column=13).value
            riesgo_personas_str = ws.cell(row=row_num, column=14).value
            observaciones = ws.cell(row=row_num, column=15).value
            persona_llamada = ws.cell(row=row_num, column=16).value
            cargo = ws.cell(row=row_num, column=17).value

            # Saltar filas vacías (sin fecha ni centro)
            if not fecha and not centro_nombre:
                filas_vacias += 1
                continue

            filas_procesadas += 1

            # --- 1) Construir fecha_hora ---
            fecha_hora = None
            if isinstance(fecha, datetime):
                if isinstance(hora, time):
                    fecha_hora = datetime.combine(fecha.date(), hora)
                else:
                    fecha_hora = fecha
            elif fecha:
                # Intentar parsear si viene como string
                try:
                    fecha_hora = datetime.strptime(str(fecha), "%Y-%m-%d %H:%M:%S")
                except:
                    self.stdout.write(self.style.WARNING(f"  Fila {row_num}: No se pudo parsear fecha '{fecha}'"))

            # --- 2) Centro ---
            centro = None
            if centro_nombre:
                centro_nombre_clean = str(centro_nombre).strip()
                centro_slug = slugify(centro_nombre_clean)
                
                if dry_run:
                    # En dry-run, solo verificamos si existe
                    try:
                        centro = Centro.objects.get(slug=centro_slug)
                    except Centro.DoesNotExist:
                        self.stdout.write(f"  [CREAR] Centro: '{centro_nombre_clean}' (slug: {centro_slug})")
                        centros_creados += 1
                else:
                    centro, created = Centro.objects.get_or_create(
                        slug=centro_slug,
                        defaults={
                            'id': centro_slug,
                            'nombre': centro_nombre_clean,
                        }
                    )
                    if created:
                        centros_creados += 1

            # --- 3) Operario ---
            operario = None
            if persona_llamada:
                persona_clean = str(persona_llamada).strip()
                cargo_clean = str(cargo).strip() if cargo else ""
                
                # Crear ID único basado en nombre + cargo + centro
                op_key = f"{persona_clean}-{cargo_clean}-{centro_nombre or ''}"
                op_id = abs(hash(op_key)) % (10**8)

                if dry_run:
                    try:
                        operario = Operario.objects.get(id=op_id)
                    except Operario.DoesNotExist:
                        self.stdout.write(f"  [CREAR] Operario: '{persona_clean}' ({cargo_clean})")
                        operarios_creados += 1
                else:
                    if centro:
                        operario, created = Operario.objects.get_or_create(
                            id=op_id,
                            defaults={
                                'nombre': persona_clean,
                                'cargo': cargo_clean,
                                'telefono': '',
                                'centro': centro,
                            }
                        )
                        if created:
                            operarios_creados += 1

            # --- 4) Convertir booleanos ---
            def to_bool(val):
                if not val:
                    return False
                v = str(val).strip().lower()
                return v in ['si', 'sí', 'yes', 'true', '1']

            riesgo_peces = to_bool(riesgo_peces_str)
            perdida_economica = to_bool(perdida_economica_str)
            riesgo_personas = to_bool(riesgo_personas_str)

            # --- 5) Limpiar strings ---
            def clean_str(val):
                return str(val).strip() if val else ""

            # Normalizar turno: convertir "Dia" a "Mañana" para consistencia
            turno_raw = clean_str(turno)
            if turno_raw.lower() in ['dia', 'día']:
                turno_clean = 'Mañana'
            else:
                turno_clean = turno_raw
            modulo_clean = clean_str(modulo)
            estanque_clean = clean_str(estanque)
            tipo_incidencia_clean = clean_str(tipo_incidencia)
            sensor_clean = clean_str(sensor)
            proveedor_clean = clean_str(proveedor)
            valor_clean = clean_str(valor)
            observaciones_clean = clean_str(observaciones)

            # Tiempo de resolución
            tiempo_res = None
            if tiempo_resolucion:
                try:
                    tiempo_res = int(tiempo_resolucion)
                except:
                    pass

            # --- 6) Crear Incidencia ---
            if dry_run:
                self.stdout.write(
                    f"  [CREAR] Incidencia: {fecha_hora} | {turno_clean} | "
                    f"{centro_nombre} | {modulo_clean} | {estanque_clean} | {tipo_incidencia_clean}"
                )
                incidencias_creadas += 1
            else:
                incidencia = Incidencia(
                    fecha_hora=fecha_hora,
                    turno=turno_clean,
                    centro=centro,
                    tipo_incidencia='modulos' if sensor_clean.lower() in ['oxígeno', 'oxigeno', 'temperatura', 'conductividad', 'turbidez'] else 'sensores',
                    modulo=modulo_clean,
                    estanque=estanque_clean,
                    parametros_afectados=sensor_clean,
                    # Determinar nivel basado en tipo_incidencia
                    oxigeno_nivel='baja' if 'bajo' in tipo_incidencia_clean.lower() else ('alta' if 'alto' in tipo_incidencia_clean.lower() else ''),
                    oxigeno_valor=valor_clean if 'ox' in sensor_clean.lower() else '',
                    temperatura_nivel='baja' if 'bajo' in tipo_incidencia_clean.lower() and 'temp' in sensor_clean.lower() else ('alta' if 'alto' in tipo_incidencia_clean.lower() and 'temp' in sensor_clean.lower() else ''),
                    temperatura_valor=valor_clean if 'temp' in sensor_clean.lower() else '',
                    sistema_sensor=proveedor_clean,
                    sensor_detectado=sensor_clean,
                    sensor_valor=valor_clean,
                    tiempo_resolucion=tiempo_res,
                    riesgo_peces=riesgo_peces,
                    perdida_economica=perdida_economica,
                    riesgo_personas=riesgo_personas,
                    observacion=observaciones_clean,
                    operario_contacto=operario,
                    tipo_incidencia_normalizada=tipo_incidencia_clean,
                )
                incidencia.save()
                incidencias_creadas += 1

        # Resumen final
        self.stdout.write("\n" + "=" * 60)
        self.stdout.write(self.style.SUCCESS(f"\n=== RESUMEN {'(DRY-RUN)' if dry_run else ''} ==="))
        self.stdout.write(f"Filas procesadas: {filas_procesadas}")
        self.stdout.write(f"Filas vacías ignoradas: {filas_vacias}")
        self.stdout.write(f"Centros {'a crear' if dry_run else 'creados'}: {centros_creados}")
        self.stdout.write(f"Operarios {'a crear' if dry_run else 'creados'}: {operarios_creados}")
        self.stdout.write(f"Incidencias {'a crear' if dry_run else 'creadas'}: {incidencias_creadas}")

        if dry_run:
            self.stdout.write(self.style.WARNING("\nEjecuta sin --dry-run para importar realmente."))
